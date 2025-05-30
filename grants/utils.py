from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from .models import GrantApplication


def export_applications_to_excel(queryset, filename_prefix="grant_applications"):
    """
    Export grant applications to Excel format
    """
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Grant Applications"
    
    # Define headers
    headers = [
        'ID', 'Applicant Name', 'Email', 'Phone', 'Status', 'Gender', 'Current Role',
        'Talk Proposal', 'Motivation', 'Contribution', 'Financial Need', 'Conference Benefit',
        'Request Travel', 'Travel From', 'Travel Amount', 'Transportation Type',
        'Request Accommodation', 'Accommodation Nights', 'Request Ticket',
        'Additional Info', 'Amount Granted', 'Reviewer Notes', 'Created At', 'Reviewed At'
    ]
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers to worksheet
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data rows
    for row_num, application in enumerate(queryset, 2):
        data = [
            application.id,
            application.user.get_full_name() or application.user.username,
            application.user.email,
            application.phone_number,
            application.get_status_display(),
            application.get_gender_display(),
            application.get_current_role_display(),
            application.get_talk_proposal_display(),
            application.motivation,
            application.contribution,
            application.financial_need,
            application.conference_benefit,
            'Yes' if application.request_travel else 'No',
            application.travel_from,
            float(application.travel_amount) if application.travel_amount else '',
            application.get_transportation_type_display() if application.transportation_type else '',
            'Yes' if application.request_accommodation else 'No',
            application.accommodation_nights if application.accommodation_nights else '',
            'Yes' if application.request_ticket else 'No',
            application.additional_info,
            float(application.amount_granted) if application.amount_granted else '',
            application.reviewer_notes,
            application.created_at.strftime('%Y-%m-%d %H:%M:%S') if application.created_at else '',
            application.reviewed_at.strftime('%Y-%m-%d %H:%M:%S') if application.reviewed_at else '',
        ]
        
        for col_num, value in enumerate(data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Set minimum width of 10 and maximum of 50
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create HTTP response
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{filename_prefix}_{timestamp}.xlsx"
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    return response


def get_applications_summary_stats(queryset):
    """
    Get summary statistics for the applications
    """
    total_applications = queryset.count()
    
    # Status breakdown
    status_stats = {}
    for status_code, status_display in GrantApplication._meta.get_field('status').choices:
        count = queryset.filter(status=status_code).count()
        status_stats[status_display] = count
    
    # Gender breakdown
    gender_stats = {}
    for gender_code, gender_display in GrantApplication._meta.get_field('gender').choices:
        count = queryset.filter(gender=gender_code).count()
        gender_stats[gender_display] = count
    
    # Request breakdown
    travel_requests = queryset.filter(request_travel=True).count()
    accommodation_requests = queryset.filter(request_accommodation=True).count()
    ticket_requests = queryset.filter(request_ticket=True).count()
    
    # Financial stats
    total_travel_amount = sum(
        app.travel_amount for app in queryset.filter(request_travel=True, travel_amount__isnull=False)
    )
    total_granted_amount = sum(
        app.amount_granted for app in queryset.filter(amount_granted__isnull=False)
    )
    
    return {
        'total_applications': total_applications,
        'status_stats': status_stats,
        'gender_stats': gender_stats,
        'travel_requests': travel_requests,
        'accommodation_requests': accommodation_requests,
        'ticket_requests': ticket_requests,
        'total_travel_amount': total_travel_amount,
        'total_granted_amount': total_granted_amount,
    } 