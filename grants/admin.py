from django.contrib import admin
from django.http import HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.utils import get_column_letter
from .models import GrantApplication


@admin.register(GrantApplication)
class GrantApplicationAdmin(admin.ModelAdmin):
    list_display = [
        'user', 'full_name', 'travel_from_display', 'travel_amount', 
        'request_travel', 'request_accommodation', 'request_ticket', 'created_at'
    ]
    list_filter = [
        'gender', 'transportation_type', 
        'request_travel', 'request_accommodation', 'request_ticket',
        'created_at'
    ]
    search_fields = [
        'user__username', 'user__email', 'user__first_name', 'user__last_name',
        'travel_from_city', 'travel_from_country'
    ]
    readonly_fields = [
        'user', 'created_at', 'updated_at'
    ]
    actions = ['export_to_excel']
    
    fieldsets = (
        ('Application Info', {
            'fields': ('user', 'created_at', 'updated_at')
        }),
        ('Personal Information', {
            'fields': ('gender', 'gender_details', 'current_role', 'current_role_details')
        }),
        ('Application Details', {
            'fields': ('motivation', 'contribution', 'financial_need')
        }),
        ('Travel Information', {
            'fields': ('request_travel', 'travel_from_city', 'travel_from_country', 'travel_amount', 'transportation_type')
        }),
        ('Accommodation & Ticket', {
            'fields': ('request_accommodation', 'accommodation_nights', 'request_ticket')
        }),
        ('Additional Information', {
            'fields': ('additional_info',)
        }),
    )
    
    def export_to_excel(self, request, queryset):
        """Export selected grant applications to Excel."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Grant Applications'
        
        headers = [
            'Username', 'First Name', 'Last Name', 'Email', 'Gender', 'Gender Details',
            'Current Role', 'Current Role Details', 'Motivation', 'Contribution', 
            'Financial Need', 'Request Travel', 'Travel From City', 'Travel From Country',
            'Travel Amount', 'Transportation Type', 'Request Accommodation', 
            'Accommodation Nights', 'Request Ticket', 'Additional Info', 'Created At'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        for row, application in enumerate(queryset.select_related('user'), 2):
            ws.cell(row=row, column=1, value=application.user.username)
            ws.cell(row=row, column=2, value=application.user.first_name)
            ws.cell(row=row, column=3, value=application.user.last_name)
            ws.cell(row=row, column=4, value=application.user.email)
            ws.cell(row=row, column=5, value=application.gender)
            ws.cell(row=row, column=6, value=application.gender_details)
            ws.cell(row=row, column=7, value=application.current_role)
            ws.cell(row=row, column=8, value=application.current_role_details)
            ws.cell(row=row, column=9, value=application.motivation)
            ws.cell(row=row, column=10, value=application.contribution)
            ws.cell(row=row, column=11, value=application.financial_need)
            ws.cell(row=row, column=12, value=application.request_travel)
            ws.cell(row=row, column=13, value=application.travel_from_city)
            ws.cell(row=row, column=14, value=application.travel_from_country.name)
            ws.cell(row=row, column=15, value=application.travel_amount)
            ws.cell(row=row, column=16, value=application.transportation_type)
            ws.cell(row=row, column=17, value=application.request_accommodation)
            ws.cell(row=row, column=18, value=application.accommodation_nights)
            ws.cell(row=row, column=19, value=application.request_ticket)
            ws.cell(row=row, column=20, value=application.additional_info)
            ws.cell(row=row, column=21, value=application.created_at.strftime('%Y-%m-%d %H:%M:%S') if application.created_at else '')
        
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 20
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="grant_applications_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
    
    export_to_excel.short_description = "Export selected applications to Excel"
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('user')
    
    def full_name(self, obj):
        return f"{obj.user.first_name} {obj.user.last_name}".strip() or obj.user.username
    full_name.short_description = 'Full Name'
    
    def travel_from_display(self, obj):
        return obj.travel_from or "Not specified"
    travel_from_display.short_description = 'Travel From'
    